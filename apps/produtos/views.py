# apps/produtos/views.py
import io
from pyexpat.errors import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    TemplateView, FormView, View
)
from datetime import date, timedelta
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.db.models import Q
from datetime import date, datetime, timedelta
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import Workbook
import pandas as pd
from apps.analytics import models
from apps.core.views import BaseMPAView
from apps.servicos.models import Servico
from .models import (
    Categoria, Fabricante,
    Produto, Lote, HistoricoPreco
)
from django.db.models import Q
from .forms import ImportarProdutosForm, LoteForm, ProdutoForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import redirect
from apps.core.models import Categoria
from .forms import CategoriaForm
from django.http import JsonResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from .models import Produto
from apps.core.models import Empresa
from django.db import models
from django.http import JsonResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.produtos.models import Produto
from apps.core.models import Empresa
from django.contrib.auth.mixins import AccessMixin
import logging
import pandas as pd
from decimal import Decimal, InvalidOperation
from django.utils import timezone
from django.core.exceptions import ValidationError
from apps.fornecedores.models import Fornecedor
import openpyxl
from openpyxl import Workbook
from apps.core.models import Categoria
import uuid
import unicodedata
from django.db import transaction
import csv
from apps.fiscal.models import TaxaIVAAGT




class PermissaoAcaoMixin(AccessMixin):
    # CRÍTICO: Definir esta variável na View
    acao_requerida = None 

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        try:
            # Tenta obter o Funcionario (ligação fundamental)
            funcionario = request.user.funcionario 
        except Exception:
            messages.error(request, "Acesso negado. O seu usuário não está ligado a um registro de funcionário.")
            return self.handle_no_permission()

        if self.acao_requerida:
            # Usa a lógica dinâmica do modelo Funcionario (que já criámos)
            if not funcionario.pode_realizar_acao(self.acao_requerida):
                messages.error(request, f"Acesso negado. O seu cargo não permite realizar a ação de '{self.acao_requerida}'.")
                return redirect(reverse_lazy('core:dashboard'))

        return super().dispatch(request, *args, **kwargs)




@login_required
def dashboard_view(request):
    """View do dashboard com notificações de estoque e validade"""
    hoje = date.today()
    data_limite = hoje + timedelta(days=30)
    
    # Lotes vencendo
    lotes_vencendo = Lote.objects.filter(
        validade__gte=hoje,
        validade__lte=data_limite
    ).select_related('produto').order_by('validade')
    
    # Produtos com estoque baixo
    produtos_ativos = Produto.objects.filter(ativo=True)
    produtos_estoque_baixo = [p for p in produtos_ativos if p.estoque_baixo]
    
    # Construir notificações
    lista_notificacoes = []
    
    for lote in lotes_vencendo:
        lista_notificacoes.append({
            'mensagem': f"Lote {lote.numero_lote} de {lote.produto.nome_produto} vence em {lote.validade.strftime('%d/%m/%Y')}",
            'detalhe': f"{lote.quantidade} unidades restantes",
            'css_class': 'bg-yellow-400'
        })

    for produto in produtos_estoque_baixo:
        lista_notificacoes.append({
            'mensagem': f"Estoque baixo: {produto.nome_produto}",
            'detalhe': f"Apenas {produto.estoque_atual} unidades",
            'css_class': 'bg-red-500'
        })
    
    context = {
        'total_notificacoes': len(lista_notificacoes),
        'lista_notificacoes': lista_notificacoes,
    }

    return render(request, 'core/dashboard.html', context)


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from .models import Produto
import json




import math
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods



@login_required
@require_http_methods(["GET"])
def buscar_produtos_api(request):
    try:
        busca = request.GET.get('search', '').strip()
        categoria_id = request.GET.get('categoria', '').strip()
        status = request.GET.get('status', '').strip()
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        
        # Empresa do usuário
        if hasattr(request.user, 'funcionario') and request.user.funcionario:
            empresa = request.user.funcionario.empresa
        else:
            return JsonResponse({'success': False, 'message': 'Empresa não encontrada'}, status=400)

        queryset = Produto.objects.filter(empresa=empresa).select_related('categoria', 'fornecedor', 'fabricante')

        if status == 'ativo':
            queryset = queryset.filter(ativo=True)
        elif status == 'inativo':
            queryset = queryset.filter(ativo=False)
        elif status == 'estoque_baixo':
            queryset = queryset.filter(estoque_atual__lte=F('estoque_minimo'))


        if categoria_id and categoria_id != 'todos':
            queryset = queryset.filter(categoria_id=categoria_id)

        if busca:
            queryset = queryset.filter(
                Q(codigo_barras__icontains=busca) |
                Q(codigo_interno__icontains=busca) |
                Q(nome_produto__icontains=busca)
            )

        total_produtos = queryset.count()  # total antes de fatiar
        start = (page - 1) * per_page
        end = start + per_page
        produtos_page = queryset[start:end]

        produtos_data = []
        for produto in produtos_page:  # <<<<<<<<<< usar produtos_page aqui
            produtos_data.append({
                'id': produto.id,
                'codigo_interno': produto.codigo_interno,
                'codigo_barras': produto.codigo_barras,
                'nome_produto': produto.nome_produto,
                'categoria': produto.categoria.nome if produto.categoria else '',
                'categoria_id': produto.categoria.id if produto.categoria else None,
                'fornecedor': str(produto.fornecedor) if produto.fornecedor else '',
                'fabricante': str(produto.fabricante) if produto.fabricante else '',
                'estoque_atual': produto.estoque_atual,
                'estoque_minimo': produto.estoque_minimo,
                'desconto_percentual': float(produto.desconto_percentual or 0),
                'preco_custo': float(produto.preco_custo or 0),
                'preco_venda': float(produto.preco_venda or 0),
                'margem_lucro': float(produto.margem_lucro or 0),
                'foto_url': produto.foto.url if produto.foto else None,
                'valor_estoque': float(produto.valor_estoque or 0),
                'estoque_baixo': produto.estoque_atual <= produto.estoque_minimo,
                'disponivel': produto.estoque_atual > 0,
                'iva_percentual': getattr(produto, 'iva_percentual', 0)
            })

        return JsonResponse({
            'success': True,
            'produtos': produtos_data,
            'total_encontrados': total_produtos,
            'page': page,
            'per_page': per_page,
            'total_paginas': math.ceil(total_produtos / per_page)
        })   

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }, status=500)
       



class ProdutosView(LoginRequiredMixin, ListView):
    template_name = 'produtos/produto_list.html'
    model = Produto
    context_object_name = 'produtos'
    paginate_by = 200

    def get_empresa(self):
        """ Método seguro para obter a empresa do utilizador logado. """
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        if user.is_superuser:
            return Empresa.objects.first()
        return None

    def get_queryset(self):
        """
        Filtra e ordena a lista de produtos.
        """
        empresa = self.get_empresa()
        if not empresa:
            return Produto.objects.none()

        # Começa com o queryset base
        queryset = Produto.objects.filter(empresa=empresa).select_related('categoria')
        
        # Filtros do URL
        search = self.request.GET.get('search', '')
        categoria_filter = self.request.GET.get('categoria', '')
        status_filter = self.request.GET.get('status', '')

        if search:
            queryset = queryset.filter(
                Q(nome_produto__icontains=search) |
                Q(codigo_barras__icontains=search)
            )
        
        if categoria_filter:
            queryset = queryset.filter(categoria_id=categoria_filter)
        
        if status_filter == 'ativo':
            queryset = queryset.filter(ativo=True)
        elif status_filter == 'inativo':
            queryset = queryset.filter(ativo=False)
        elif status_filter == 'estoque_baixo':
            queryset = queryset.filter(estoque_atual__lte=F('estoque_minimo'))
        
        return queryset.order_by('nome_produto')

    def get_context_data(self, **kwargs):
        """ Adiciona dados extra (estatísticas, etc.) ao contexto. """
        context = super().get_context_data(**kwargs)
        empresa = self.get_empresa()

        if empresa:
            todas_categorias = Categoria.objects.filter(empresa=empresa).order_by('nome')
            servicos = Servico.objects.filter(empresa=empresa)
            todos_produtos = Produto.objects.filter(empresa=empresa)
            
            context['produtos_stats'] = {
                'total': todos_produtos.count(),
                'ativos': todos_produtos.filter(ativo=True).count(),
                'categorias': todas_categorias.count(),
                'servicos': servicos.count(),
            }
            context['categorias'] = todas_categorias.filter(ativa=True)
        
        return context




class CriarProdutoView(LoginRequiredMixin, CreateView):
    model = Produto
    form_class = ProdutoForm
    template_name = 'produtos/produto_form.html'
    success_url = reverse_lazy('produtos:produto_list')
    acao_requerida = 'editar_produtos'

    def get_empresa(self):
        """ Método seguro para obter a empresa do utilizador logado. """
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        if user.is_superuser:
            from apps.core.models import Empresa
            return Empresa.objects.first()
        return None

    def dispatch(self, request, *args, **kwargs):
        """ Verifica se o utilizador tem uma empresa antes de continuar. """
        if not self.get_empresa():
            messages.error(request, "O seu utilizador não está associado a nenhuma empresa.")
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        """ Passa a empresa para o formulário para filtrar os ForeignKeys. """
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa()
        return kwargs

    def form_valid(self, form):
        produto = form.save(commit=False)
        produto.empresa = self.get_empresa()
        produto.save()
        messages.success(self.request, f"Produto '{produto.nome_produto}' criado com sucesso!")
        
        # MELHORIA: Deixe a CreateView tratar do redirecionamento
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Adicionar Novo Produto"
        return context


class EditarProdutoView(LoginRequiredMixin, UpdateView):
    model = Produto
    form_class = ProdutoForm
    template_name = 'produtos/produto_form.html'
    success_url = reverse_lazy('produtos:produto_list') # Redireciona para a lista de produtos
    pk_url_kwarg = 'produto_id' # Informa que o ID na URL é 'produto_id'
    acao_requerida = 'editar_produtos'

    def get_queryset(self):
        """ Garante que o utilizador só pode editar produtos da sua própria empresa. """
        empresa = self.request.user.funcionario.empresa
        return Produto.objects.filter(empresa=empresa)

    def get_form_kwargs(self):
        """ Passa a empresa para o formulário para filtrar os campos ForeignKey. """
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.request.user.funcionario.empresa
        return kwargs

    def form_valid(self, form):
        """ Adiciona uma mensagem de sucesso antes de redirecionar. """
        messages.success(self.request, f"Produto '{form.instance.nome_produto}' atualizado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Editar Produto: {self.object.nome_produto}"
        return context
    

class DeletarProdutoView(LoginRequiredMixin, View):
    acao_requerida = 'editar_produtos'
    def post(self, request, produto_id):
        try:
            empresa = self.get_empresa(request)
            from apps.produtos.models import Produto
            produto = Produto.objects.get(id=produto_id, empresa=empresa)
            
            nome = produto.nome_produto
            produto.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Produto "{nome}" removido com sucesso'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    def get_empresa(self, request):
        if hasattr(request.user, 'usuario') and request.user.usuario.empresa:
            return request.user.usuario.empresa
        elif hasattr(request.user, 'profile') and request.user.profile.empresa:
            return request.user.profile.empresa
        else:
            from apps.core.models import Empresa
            return Empresa.objects.first()


class DeletarTodosProdutosLotesView(LoginRequiredMixin, View):
    """
    Remove TODOS os produtos e TODOS os lotes.
    Uso: Ação drástica de limpeza.
    """
    def post(self, request):
        try:
            empresa = self._get_empresa(request)
            from apps.produtos.models import Produto
            
            # Como on_delete=CASCADE no Lote, apagar produtos apaga lotes
            count, _ = Produto.objects.filter(empresa=empresa).delete()
            
            return JsonResponse({
                'success': True,
                'message': f'{count} registos (produtos e dependências) removidos permanentemente.'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    def _get_empresa(self, request):
        # Helper simples
        if hasattr(request.user, 'usuario') and request.user.usuario.empresa:
            return request.user.usuario.empresa
        return None


class DeletarTodosProdutosManterLotesView(LoginRequiredMixin, View):
    """
    Remove (Desativa) TODOS os produtos, mantendo os lotes no banco.
    Nota: Como Lote tem FK(Produto, on_delete=CASCADE), não podemos apagar o produto fisicamente
    sem apagar o lote. Portanto, esta view faz um SOFT DELETE (ativo=False).
    """
    def post(self, request):
        try:
            empresa = self._get_empresa(request)
            from apps.produtos.models import Produto
            
            # Soft Delete
            updated = Produto.objects.filter(empresa=empresa, ativo=True).update(ativo=False)
            
            return JsonResponse({
                'success': True,
                'message': f'{updated} produtos foram desativados. Os lotes permanecem inalterados.'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    def _get_empresa(self, request):
        if hasattr(request.user, 'usuario') and request.user.usuario.empresa:
            return request.user.usuario.empresa
        return None


class ToggleProdutoView(LoginRequiredMixin, View):
    acao_requerida = 'editar_produtos'
    def post(self, request, produto_id):
        try:
            empresa = self.get_empresa(request)
            from apps.produtos.models import Produto
            produto = Produto.objects.get(id=produto_id, empresa=empresa)
            
            produto.ativo = not produto.ativo
            produto.save()
            
            return JsonResponse({
                'success': True,
                'ativo': produto.ativo,
                'message': f'Produto {"ativado" if produto.ativo else "desativado"} com sucesso'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    def get_empresa(self, request):
        if hasattr(request.user, 'usuario') and request.user.usuario.empresa:
            return request.user.usuario.empresa
        elif hasattr(request.user, 'profile') and request.user.profile.empresa:
            return request.user.profile.empresa
        else:
            from apps.core.models import Empresa
            return Empresa.objects.first()


# apps/produtos/views.py (substituir a TemplateProdutosView existente)

class CategoriaBaseView(LoginRequiredMixin):
    """
    View base para garantir que as operações são feitas na empresa correta.
    Agora com métodos seguros para obter a empresa.
    """
    model = Categoria

    def get_empresa(self):
        """ Método seguro para obter a empresa do utilizador logado. """
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        # Adicione um fallback para o superuser ou outros casos, se necessário
        if user.is_superuser and not hasattr(user, 'funcionario'):
             from apps.core.models import Empresa
             return Empresa.objects.first()
        return None

    def dispatch(self, request, *args, **kwargs):
        """ Garante que o utilizador tem uma empresa antes de prosseguir. """
        if not self.get_empresa():
            messages.error(request, "O seu utilizador não está associado a nenhuma empresa.")
            return redirect('core:dashboard') # Redireciona para o dashboard principal
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        """ Filtra as categorias para mostrar apenas as da empresa do utilizador. """
        return Categoria.objects.filter(empresa=self.get_empresa())

class CategoriaListView(CategoriaBaseView, ListView):
    template_name = 'produtos/categoria_list.html'
    context_object_name = 'categorias'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Categorias de Produtos"
        return context



class CategoriaCreateView(CategoriaBaseView, CreateView):
    model = Categoria
    # --- PONTO CRÍTICO DA CORREÇÃO ---
    form_class = CategoriaForm
    
    template_name = 'produtos/categoria_form.html'
    success_url = reverse_lazy('produtos:categoria_list')

    def get_form_kwargs(self):
        # Este método é essencial para passar a 'empresa' para o formulário
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa()
        return kwargs

    def form_valid(self, form):
        # A lógica aqui está correta, associa a empresa antes de salvar
        form.instance.empresa = self.get_empresa()
        messages.success(self.request, "Categoria criada com sucesso!")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Adicionar Nova Categoria"
        return context       



class CategoriaUpdateView(CategoriaBaseView, UpdateView):
    form_class = CategoriaForm
    template_name = 'produtos/categoria_form.html'
    success_url = reverse_lazy('produtos:categoria_list')

    def get_form_kwargs(self):
        """ Passa a empresa para o formulário. """
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa()
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Categoria atualizada com sucesso!")
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Editar Categoria: {self.object.nome}"
        return context

class CategoriaDeleteView(CategoriaBaseView, DeleteView):
    template_name = 'produtos/categoria_confirm_delete.html'
    success_url = reverse_lazy('produtos:categoria_list')
    
    def post(self, request, *args, **kwargs):
        # O método post é o correto para adicionar a mensagem antes de apagar
        self.object = self.get_object()
        messages.success(self.request, f"Categoria '{self.object.nome}' eliminada com sucesso!")
        return self.delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Eliminar Categoria: {self.object.nome}"
        return context   



#######################################



# ===============================
# LOTE VIEWS
# ===============================

class LoteListView(LoginRequiredMixin, ListView):
    model = Lote
    template_name = "produtos/lote_list.html"
    context_object_name = "lotes"
    paginate_by = 200  # Define a paginação aqui

    def get_queryset(self):
        """
        Este método agora constrói a consulta ao banco de dados de forma dinâmica,
        aplicando os filtros e a ordenação recebidos do formulário.
        """
        # Começa com todos os lotes (respeitando permissões, se aplicável)
        queryset = Lote.objects.select_related('produto').all()

        # Obtém os parâmetros do URL (ex: ?search=paracetamol&status=vencendo)
        search_query = self.request.GET.get('search', '').strip()
        status_filter = self.request.GET.get('status', '')
        sort_by = self.request.GET.get('sort', '-data_validade') # Padrão: validade mais longe

        # 1. Aplicar filtro de pesquisa
        if search_query:
            queryset = queryset.filter(
                Q(produto__nome_produto__icontains=search_query) |
                Q(numero_lote__icontains=search_query)
            )

        # 2. Aplicar filtro de status de validade
        hoje = date.today()
        if status_filter == 'vencido':
            queryset = queryset.filter(data_validade__lt=hoje)
        elif status_filter == 'vencendo':
            data_limite = hoje + timedelta(days=30)
            queryset = queryset.filter(data_validade__gte=hoje, data_validade__lte=data_limite)

        # 3. Aplicar ordenação
        # Validar as opções de ordenação para segurança
        valid_sort_options = ['data_validade', '-data_validade', 'quantidade_atual', '-quantidade_atual']
        if sort_by in valid_sort_options:
            queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        """ Adiciona os valores dos filtros ao contexto para manter o estado no template. """
        context = super().get_context_data(**kwargs)
        # Passa os valores atuais dos filtros de volta para o template
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_sort'] = self.request.GET.get('sort', '-data_validade')
        context['add_product_form'] = ProdutoForm() 
        
        
        return context


class LoteDetailView(LoginRequiredMixin, DetailView):
    model = Lote
    template_name = "produtos/lote_detail.html"






from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView

from .models import Lote
from .forms import LoteForm, ProdutoForm # Garanta que os forms estão importados

# --- BASE VIEW PARA LOTES (BOA PRÁTICA) ---
class LoteBaseView(LoginRequiredMixin):
    """ View base para Lotes, garante o modelo e o filtro por empresa. """
    model = Lote

    def get_queryset(self):
        # Garante que qualquer view que herde disto só aceda a lotes da empresa do utilizador
        empresa = self.request.user.funcionario.empresa
        return Lote.objects.filter(produto__empresa=empresa)

# --- CRUD COMPLETO PARA LOTES ---

class LoteListView(LoteBaseView, ListView):
    template_name = "produtos/lote_list.html"
    context_object_name = "lotes"
    paginate_by = 20

    def get_queryset(self):
        # Começa com o queryset já filtrado por empresa da LoteBaseView
        queryset = super().get_queryset().select_related('produto')
        
        search_query = self.request.GET.get('search', '').strip()
        status_filter = self.request.GET.get('status', '')
        sort_by = self.request.GET.get('sort', '-data_validade')

        if search_query:
            # CORREÇÃO: Usar 'nome_comercial'
            queryset = queryset.filter(
                Q(produto__nome_comercial__icontains=search_query) |
                Q(numero_lote__icontains=search_query)
            )

        hoje = date.today()
        if status_filter == 'vencido':
            queryset = queryset.filter(data_validade__lt=hoje)
        elif status_filter == 'vencendo':
            data_limite = hoje + timedelta(days=30)
            queryset = queryset.filter(data_validade__gte=hoje, data_validade__lte=data_limite)

        valid_sort_options = ['data_validade', '-data_validade', 'quantidade_atual', '-quantidade_atual']
        if sort_by in valid_sort_options:
            queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_sort'] = self.request.GET.get('sort', '-data_validade')
        context['add_product_form'] = ProdutoForm() 
        context['title'] = "Gestão de Lotes de Produtos"
        return context

class LoteDetailView(LoteBaseView, DetailView):
    template_name = "produtos/lote_detail.html"
    context_object_name = "lote"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalhes do Lote: {self.object.numero_lote}"
        return context

class LoteCreateView(LoteBaseView, CreateView):
    form_class = LoteForm
    template_name = "produtos/lote_form.html"
    success_url = reverse_lazy("produtos:lote_list")

    def form_valid(self, form):
        lote = form.save(commit=False)
        lote.quantidade_atual = form.cleaned_data['quantidade_inicial']
        lote.save()
        messages.success(self.request, f"Lote {lote.numero_lote} adicionado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Adicionar Novo Lote"
        return context

class LoteUpdateView(LoteBaseView, UpdateView):
    form_class = LoteForm
    template_name = "produtos/lote_form.html"
    
    def get_success_url(self):
        return reverse_lazy('produtos:lote_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Lote atualizado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Editar Lote: {self.object.numero_lote}"
        return context

class LoteDeleteView(LoteBaseView, DeleteView):
    template_name = "produtos/lote_confirm_delete.html"
    success_url = reverse_lazy("produtos:lote_list")
    
    def form_valid(self, form):
        messages.success(self.request, f"Lote '{self.object.numero_lote}' eliminado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Eliminar Lote"
        return context
    
# ===============================
# AJAX VIEWS
# ===============================

class BuscarProdutoAjaxView(LoginRequiredMixin, View):
    def get(self, request):
        termo = request.GET.get('q', '').strip()
        
        if len(termo) < 2:
            return JsonResponse({'produtos': []})
        
        produtos = Produto.objects.filter(
            Q(nome_produto__icontains=termo) |
            Q(codigo_barras__icontains=termo) |
            Q(codigo_interno__icontains=termo),
            ativo=True
        ).select_related('categoria', 'fabricante')[:10]
        
        produtos_data = []
        for produto in produtos:
            produtos_data.append({
                'id': produto.id,
                'nome_produto': produto.nome_produto,
                'codigo_barras': produto.codigo_barras,
                'preco_venda': float(produto.preco_venda),
                'fabricante': produto.fabricante.nome if produto.fabricante else '',
                'categoria': produto.categoria.nome if produto.categoria else '',
                'estoque_atual': produto.estoque_atual,
            })
        
        return JsonResponse({'produtos': produtos_data})


class VerificarCodigoView(LoginRequiredMixin, View):
    def get(self, request):
        codigo = request.GET.get('codigo', '').strip()
        existe = Produto.objects.filter(codigo_barras=codigo).exists()
        return JsonResponse({'existe': existe})

# ===============================
# RELATÓRIOS E CONSULTAS
# ===============================

class LotesVencimentoView(LoginRequiredMixin, ListView):
    model = Lote
    template_name = "produtos/lotes_vencimento.html"
    context_object_name = "lotes"

    def get_queryset(self):
        hoje = timezone.now().date()
        dias = int(self.request.GET.get('dias', 30))
        data_limite = hoje + timedelta(days=dias)
        
        return Lote.objects.filter(
            validade__gte=hoje,
            validade__lte=data_limite,
            quantidade__gt=0
        ).select_related('produto').order_by('validade')

class ProdutosEstoqueBaixoView(LoginRequiredMixin, ListView):
    model = Produto
    template_name = "produtos/estoque_baixo.html"
    context_object_name = "produtos"

    def get_queryset(self):
        produtos = Produto.objects.filter(ativo=True, estoque_minimo__gt=0)
        produtos_baixo = []
        
        for produto in produtos:
            if produto.estoque_baixo:
                produtos_baixo.append(produto)
        
        return produtos_baixo


# ===============================
# AÇÕES ESPECIAIS
# ===============================

class AtivarProdutoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        produto = get_object_or_404(Produto, pk=pk)
        produto.ativo = True
        produto.save()
        return redirect("produtos:produto_detail", pk=pk)

class DesativarProdutoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        produto = get_object_or_404(Produto, pk=pk)
        produto.ativo = False
        produto.save()
        return redirect("produtos:produto_detail", pk=pk)


@login_required
@require_http_methods(["GET"])
def listar_categorias_api(request):
    """
    API para listar categorias
    """
    try:
        if hasattr(request.user, 'funcionario') and request.user.funcionario:
            empresa = request.user.funcionario.empresa
        else:
            return JsonResponse({
                'success': False,
                'message': 'Empresa não encontrada'
            }, status=400)
        
        from .models import Categoria
        categorias = Categoria.objects.filter(empresa=empresa, ativo=True)
        
        categorias_data = []
        for categoria in categorias:
            categorias_data.append({
                'id': categoria.id,
                'nome': categoria.nome,
                'descricao': getattr(categoria, 'descricao', ''),
            })
        
        return JsonResponse({
            'success': True,
            'categorias': categorias_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }, status=500)

from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Fabricante
from .forms import FabricanteForm

# Lista de fabricantes
class FabricanteListView(LoginRequiredMixin, ListView):
    model = Fabricante
    template_name = 'fabricantes/fabricante_lista.html'
    context_object_name = 'fabricantes'
    paginate_by = 25  # opcional
    ordering = ['nome']

    def get_queryset(self):
        return Fabricante.objects.filter(empresa=self.request.user.funcionario.empresa)

# Criar fabricante
class FabricanteCreateView(LoginRequiredMixin, CreateView):
    model = Fabricante
    form_class = FabricanteForm
    template_name = 'fabricantes/form.html'
    success_url = reverse_lazy('fabricantes:fabricante_lista')

    def form_valid(self, form):
        form.instance.empresa = self.request.user.funcionario.empresa
        return super().form_valid(form)

# Editar fabricante
class FabricanteUpdateView(LoginRequiredMixin, UpdateView):
    model = Fabricante
    form_class = FabricanteForm
    template_name = 'fabricantes/form.html'
    success_url = reverse_lazy('fabricantes:fabricante_lista')

# Detalhes do fabricante
class FabricanteDetailView(LoginRequiredMixin, DetailView):
    model = Fabricante
    template_name = 'fabricantes/detail.html'
    context_object_name = 'fabricante'

# Excluir fabricante
class FabricanteDeleteView(LoginRequiredMixin, DeleteView):
    model = Fabricante
    template_name = 'fabricantes/confirm_delete.html'
    success_url = reverse_lazy('fabricantes:fabricante_lista')



class TemplateProdutosView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            empresa = self.get_empresa()
            if not empresa:
                return HttpResponseBadRequest('Empresa não encontrada')
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"
            
            # Definir cabeçalhos
            headers = [
                'nome_produto', 'codigo_barras', 'categoria', 'preco_custo', 'preco_venda', 'taxa',
                'nome_comercial', 'codigo_interno', 'fabricante', 'fornecedor',
                'numero_lote', 'data_validade',
                'estoque_atual', 'estoque_minimo', 'estoque_maximo', 'margem_lucro',
                'desconto_percentual', 'observacoes', 'ativo'
            ]
            
            # Adicionar cabeçalhos
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Adicionar dados de exemplo
            exemplo = [
                'Paracetamol 500mg',    # nome_produto
                '7891234567890',        # codigo_barras
                'Analgésicos',         # categoria
                '105.50',                 # preco_custo
                '200.00',                 # preco_venda
                '14',                   # taxa
                'Paracetamol 500mg',    # nome_comercial
                'PARA500',              # codigo_interno
                'NEWAY FARMÁCIA',                  # fabricante
                'NEWAY MED',    # fornecedor
                'LOTE202401',           # numero_lote
                '2025-12-31',           # data_validade
                '100',                  # estoque_atual
                '10',                   # estoque_minimo
                '1000',                 # estoque_maximo
                '100.00',               # margem_lucro
                '0.00',                 # desconto_percentual
                'Produto de template',   # observacoes
                '1'                     # ativo
            ]
            
            for col, value in enumerate(exemplo, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Adicionar instruções em uma nova planilha
            ws_instrucoes = wb.create_sheet("Instruções")
            
            instrucoes = [
                "INSTRUÇÕES PARA IMPORTAÇÃO DE PRODUTOS",
                "",
                "CAMPOS OBRIGATÓRIOS:",
                "- nome_produto: Nome do produto",
                "- codigo_barras: Código de barras único",
                "- categoria: Nome da categoria",
                "- preco_custo: Preço de custo (formato: 12.50)",
                "- preco_venda: Preço de venda (formato: 25.00)",
                "",
                "CAMPOS OPCIONAIS (MAS RECOMENDADOS):",
                "- taxa: Percentagem de IVA (ex: 14, 7, 0). Padrão: 0. Se não existir, será criada.",
                "- nome_comercial: Nome comercial (se vazio, usará nome_produto)",
                "- codigo_interno: Código interno (se vazio, usará codigo_barras)",
                "- fabricante: Nome do fabricante",
                "- fornecedor: Nome do fornecedor (deve existir no sistema)",
                "- numero_lote: Número do lote (se fornecido, um lote será criado/atualizado)",
                "- data_validade: Validade do lote (fomato YYYY-MM-DD ou DD/MM/YYYY)",
                "- estoque_atual: Quantidade em estoque (padrão: 0)",
                "- estoque_minimo: Estoque mínimo (padrão: 1)",
                "- estoque_maximo: Estoque máximo (padrão: 100)",
                "- margem_lucro: Margem de lucro em % (padrão: 0)",
                "- desconto_percentual: Desconto em % (padrão: 0)",
                "- observacoes: Observações sobre o produto",
                "- ativo: 1 para ativo, 0 para inativo (padrão: 1)",
                "",
                "FORMATOS:",
                "- Preços/Taxas: Use ponto como separador decimal (ex: 12.50)",
                "- Datas: Formato AAAA-MM-DD é o mais seguro para CSV",
                "- Booleanos: Use 1/0 ou true/false ou sim/não",
                "- Texto: Evite caracteres especiais",
                "",
                "NOTAS IMPORTANTES:",
                "ATT: SUPORTE PARA ATÉ 1000 LINHAS DE PRODUTOS SIMULTANEAENTE",
                "- Códigos de barras devem ser únicos na empresa",
                "- Categorias, Fabricantes e Taxas de IVA serão criados automaticamente se não existirem",
                "- Fornecedores devem existir previamente no sistema",
                "- Se informar Lote e Validade, o sistema atualizará o estoque desse lote específico",
                "- Linhas com erros serão registradas no log, mas o processo tentará continuar",
                "- Use a opção 'Validar apenas' para testar antes de importar"
            ]
            
            for row, instrucao in enumerate(instrucoes, 1):
                ws_instrucoes.cell(row=row, column=1, value=instrucao)
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Salvar em buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Retornar response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="template_produtos_{empresa.nome}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            
            return response
            
        except Exception as e:
            logger.error(f"Erro ao gerar template: {str(e)}")
            return HttpResponseBadRequest(f'Erro ao gerar template: {str(e)}')

    def get_empresa(self):
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario and user.funcionario.empresa:
            return user.funcionario.empresa
        if user.is_superuser:
            return Empresa.objects.first()
        return None

logger = logging.getLogger(__name__)




class ImportarProdutosView(LoginRequiredMixin, FormView):
    template_name = 'produtos/importar_produtos.html'
    form_class = ImportarProdutosForm
    success_url = reverse_lazy('produtos:produto_list')
    acao_requerida = 'editar_produtos'

    # ------------------------------
    # Utilidades auxiliares
    # ------------------------------
    def get_empresa(self):
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario.empresa:
            return user.funcionario.empresa
        return None

    def sanitize_text(self, text):
        if not text:
            return ''
        try:
            text = str(text)
            text = unicodedata.normalize('NFKC', text)
            return text.strip()
        except Exception:
            return str(text).strip()

    def obter_categoria(self, nome, empresa):
        # NOTE: Deprecated in favor of bulk fetching inside processar_arquivo
        return None

    def obter_fornecedor(self, nome, empresa):
        # NOTE: Deprecated in favor of bulk fetching inside processar_arquivo
        return None

    def obter_fabricante(self, nome, empresa):
        # NOTE: Deprecated in favor of bulk fetching inside processar_arquivo
        return None

    # ------------------------------
    # Processamento principal
    # ------------------------------
    def processar_arquivo(self, arquivo, empresa):
        registros_criados, registros_atualizados = 0, 0
        
        # --- Helper interno ---
        def get_val(row, col, default=None):
            val = row.get(col, default)
            # Handle pandas Series/Objects
            if hasattr(val, 'iloc'):
                val = val.iloc[0]
            elif isinstance(val, list):
                val = val[0] if val else default
            
            return val if pd.notna(val) else default

        try:
            conteudo = arquivo.read()
            nome_ext = arquivo.name.lower()

            # 1. Carregar DataFrame
            if nome_ext.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(io.BytesIO(conteudo))
            else:
                try:
                    texto = conteudo.decode('utf-8')
                except UnicodeDecodeError:
                    texto = conteudo.decode('latin1', errors='ignore')

                try:
                    sniffer = csv.Sniffer()
                    sep = sniffer.sniff(texto.splitlines()[0]).delimiter
                except Exception:
                    sep = ';' if texto.count(';') > texto.count(',') else ','
                df = pd.read_csv(io.StringIO(texto), sep=sep, on_bad_lines='skip')

            if df.empty:
                logger.warning("O arquivo foi lido mas não contém dados!")
                return 0, 0

            # 2. Normalizar Colunas
            df.columns = [str(c).strip().lower() for c in df.columns]
            df = df.loc[:, ~df.columns.duplicated()]

            col_map = {
                'nome_produto': 'nome', 'produto': 'nome', 'nome': 'nome',
                'descricao': 'nome', 'descrição': 'nome',
                'preco': 'preco_venda', 'preço': 'preco_venda', 'valor': 'preco_venda',
                'preco_custo': 'preco_custo', 'preco_venda': 'preco_venda',
                'taxa': 'taxa', 'iva': 'taxa', 'imposto': 'taxa', 'percentage': 'taxa',
                'categoria': 'categoria', 'grupo': 'categoria',
                'fornecedor': 'fornecedor', 'fabricante': 'fabricante',
                'codigo_barras': 'codigo_barras', 'codigo_interno': 'codigo_interno',
                'nome_comercial': 'nome_comercial',
                'numero_lote': 'numero_lote', 'lote': 'numero_lote',
                'validade': 'data_validade', 'data_validade': 'data_validade', 'vencimento': 'data_validade',
                'estoque_atual': 'estoque_atual', 'estoque': 'estoque_atual', 
                'estoque_minimo': 'estoque_minimo', 'estoque_maximo': 'estoque_maximo', 
                'margem_lucro': 'margem_lucro', 'desconto_percentual': 'desconto_percentual',
                'observacoes': 'observacoes', 'ativo': 'ativo',
            }
            df.rename(columns=lambda x: col_map.get(x, x), inplace=True)
            logger.info(f"Colunas mapeadas: {list(df.columns)}")

            # 3. Otimização: Pre-fetching de Dependências (Bulk)
            
            # --- Taxas IVA (bulk) ---
            if 'taxa' in df.columns:
                # Converter para numérico e preencher vazios com 0
                df['taxa'] = pd.to_numeric(df['taxa'], errors='coerce').fillna(0)
            else:
                df['taxa'] = 0

            taxas_unicas = df['taxa'].unique()
            # Precisamos normalizar para Decimal para comparar com o banco
            taxas_decimals = set()
            for t in taxas_unicas:
                try:
                    taxas_decimals.add(Decimal(str(t)).quantize(Decimal('0.00')))
                except:
                    taxas_decimals.add(Decimal('0.00'))
            
            # Buscar existentes
            taxas_existentes = TaxaIVAAGT.objects.filter(
                empresa=empresa,
                tax_percentage__in=taxas_decimals
            )
            # Map: {Decimal('14.00'): TaxaIVAAGT(...)}
            taxa_map = {t.tax_percentage: t for t in taxas_existentes}
            
            # Criar novas taxas
            novas_taxas = []
            for tax_val in taxas_decimals:
                if tax_val not in taxa_map:
                     novas_taxas.append(TaxaIVAAGT(
                        empresa=empresa,
                        nome=f"IVA {tax_val}%",
                        tax_type='IVA',
                        tax_code='NOR',
                        tax_percentage=tax_val,
                        exemption_reason=None if tax_val > 0 else 'M99', # Se 0, precisa motivo, mas vamos null ou M99 genérico
                        ativo=True
                     ))
            
            if novas_taxas:
                TaxaIVAAGT.objects.bulk_create(novas_taxas)
                # Refetch para garantir integridade e pegar IDs se necessário
                all_taxes = TaxaIVAAGT.objects.filter(empresa=empresa, tax_percentage__in=taxas_decimals)
                taxa_map = {t.tax_percentage: t for t in all_taxes}

            # --- Categorias ---
            nomes_cat = df['categoria'].dropna().unique()
            nomes_cat_limpos = {self.sanitize_text(n) for n in nomes_cat if self.sanitize_text(n)}
            
            # Buscar existentes
            cats_existentes = Categoria.objects.filter(
                empresa=empresa, nome__in=nomes_cat_limpos
            ).values_list('nome', 'id')
            cat_map = {nome.lower(): pk for nome, pk in cats_existentes}

            # Criar novas categorias
            novas_cats = []
            for nome in nomes_cat_limpos:
                if nome.lower() not in cat_map:
                    novas_cats.append(Categoria(empresa=empresa, nome=nome, ativa=True))
            
            if novas_cats:
                Categoria.objects.bulk_create(novas_cats)
                # Recarregar mapa
                cats_all = Categoria.objects.filter(empresa=empresa, nome__in=nomes_cat_limpos)
                cat_map = {c.nome.lower(): c for c in cats_all}
            else:
                # Recarregar objetos completos para o mapa
                cats_all = Categoria.objects.filter(empresa=empresa, nome__in=nomes_cat_limpos)
                cat_map = {c.nome.lower(): c for c in cats_all}

            # --- Fabricantes ---
            nomes_fab = df['fabricante'].dropna().unique()
            nomes_fab_limpos = {self.sanitize_text(n) for n in nomes_fab if self.sanitize_text(n)}
            
            fabs_existentes = Fabricante.objects.filter(
                empresa=empresa, nome__in=nomes_fab_limpos
            ).values('nome')
            fab_existem_set = {f['nome'].lower() for f in fabs_existentes}
            
            novos_fabs = []
            for nome in nomes_fab_limpos:
                if nome.lower() not in fab_existem_set:
                    novos_fabs.append(Fabricante(empresa=empresa, nome=nome, ativo=True))
            
            if novos_fabs:
                Fabricante.objects.bulk_create(novos_fabs)
            
            fabs_all = Fabricante.objects.filter(empresa=empresa, nome__in=nomes_fab_limpos)
            fab_map = {f.nome.lower(): f for f in fabs_all}

            # --- Fornecedores ---
            # Fornecedor não criamos automaticamente por segurança, apenas buscamos
            nomes_forn = df['fornecedor'].dropna().unique()
            nomes_forn_limpos = {self.sanitize_text(n) for n in nomes_forn if self.sanitize_text(n)}
            
            forn_all = Fornecedor.objects.filter(
                empresa=empresa
            ).filter(
                Q(nome_fantasia__in=nomes_forn_limpos) | Q(razao_social__in=nomes_forn_limpos)
            )
            # Mapper mais complexo: tenta nome fantasia e razao social
            forn_map = {} # chave: nome limpo lower -> objeto
            for f in forn_all:
                if f.nome_fantasia: forn_map[f.nome_fantasia.strip().lower()] = f
                if f.razao_social: forn_map[f.razao_social.strip().lower()] = f

            # 4. Processamento Transacional
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        nome = self.sanitize_text(get_val(row, 'nome'))
                        if not nome:
                            continue

                        # Resolve FKs via Mapas
                        cat_nome = self.sanitize_text(get_val(row, 'categoria'))
                        categoria = cat_map.get(cat_nome.lower()) if cat_nome else None

                        fab_nome = self.sanitize_text(get_val(row, 'fabricante'))
                        fabricante = fab_map.get(fab_nome.lower()) if fab_nome else None

                        forn_nome = self.sanitize_text(get_val(row, 'fornecedor'))
                        fornecedor = forn_map.get(forn_nome.lower()) if forn_nome else None

                        # Resolve Taxa IVA
                        # Obter valor e converter para Decimal para lookup
                        def clean_decimal(v, default=0):
                             if pd.isna(v) or v == '': return Decimal(default)
                             try: return Decimal(str(v).replace(',', '.'))
                             except: return Decimal(default)

                        taxa_val_raw = clean_decimal(get_val(row, 'taxa'), 0)
                        taxa_val_key = taxa_val_raw.quantize(Decimal('0.00')) # Chave do mapa
                        # Fallback seguro para 0 se não achar (embora deva achar pois criamos todas)
                        taxa_iva = taxa_map.get(taxa_val_key) or taxa_map.get(Decimal('0.00'))

                        # Valores básicos
                        codigo_barras = self.sanitize_text(get_val(row, 'codigo_barras'))
                        if not codigo_barras:
                             # Se não tem codigo de barras, geramos um temporário ou ignoramos?
                             # Regra original: gerava UUID
                             codigo_barras = f"AUTO-{uuid.uuid4().hex[:10]}"
                        
                        codigo_interno = self.sanitize_text(get_val(row, 'codigo_interno')) or f"INT-{uuid.uuid4().hex[:8]}"
                        nome_comercial = self.sanitize_text(get_val(row, 'nome_comercial')) or nome

                        # Numéricos
                        preco_custo = clean_decimal(get_val(row, 'preco_custo'), 0)
                        preco_venda = clean_decimal(get_val(row, 'preco_venda'), 0)
                        estoque_atual = clean_decimal(get_val(row, 'estoque_atual'), 0)
                        estoque_minimo = clean_decimal(get_val(row, 'estoque_minimo'), 1)
                        estoque_maximo = clean_decimal(get_val(row, 'estoque_maximo'), 100)
                        margem_lucro = clean_decimal(get_val(row, 'margem_lucro'), 0)
                        desconto_percentual = clean_decimal(get_val(row, 'desconto_percentual'), 0)
                        
                        observacoes = self.sanitize_text(get_val(row, 'observacoes'))
                        
                        val_ativo = get_val(row, 'ativo', 1)
                        ativo = str(val_ativo).lower() in ['1', 'true', 'sim', 'yes', 'verdadeiro']

                        # --- Produto: Update ou Create ---
                        produto, created = Produto.objects.update_or_create(
                            empresa=empresa,
                            codigo_barras=codigo_barras,
                            defaults={
                                'codigo_interno': codigo_interno,
                                'nome_produto': nome,
                                'nome_comercial': nome_comercial,
                                'categoria': categoria,
                                'fornecedor': fornecedor,
                                'fabricante': fabricante,
                                'taxa_iva': taxa_iva,
                                'preco_custo': preco_custo,
                                'preco_venda': preco_venda,
                                'estoque_atual': estoque_atual,
                                'estoque_minimo': estoque_minimo,
                                'estoque_maximo': estoque_maximo,
                                'margem_lucro': margem_lucro,
                                'desconto_percentual': desconto_percentual,
                                'observacoes': observacoes,
                                'ativo': ativo,
                            }
                        )

                        if created:
                            registros_criados += 1
                        else:
                            registros_atualizados += 1

                        # --- Lote / Stock ---
                        numero_lote = self.sanitize_text(get_val(row, 'numero_lote'))
                        data_validade_raw = get_val(row, 'data_validade')

                        if numero_lote and data_validade_raw:
                            # Tentar parse da data
                            data_validade = None
                            if hasattr(data_validade_raw, 'date'): # Se for datetime/timestamp do pandas
                                data_validade = data_validade_raw.date()
                            else:
                                try:
                                    # Tenta formatos comuns
                                    d_str = str(data_validade_raw).strip()
                                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
                                        try:
                                            data_validade = datetime.strptime(d_str, fmt).date()
                                            break
                                        except ValueError:
                                            continue
                                except Exception:
                                    pass

                            if data_validade:
                                # Criar ou atualizar lote
                                # Se o ficheiro diz estoque_atual = 50, assumimos que esse lote tem 50?
                                # A lógica original usava 'estoque_atual' para o produto.
                                # Vamos atribuir esse estoque a este lote.
                                
                                Lote.objects.update_or_create(
                                    produto=produto,
                                    numero_lote=numero_lote,
                                    defaults={
                                        'data_validade': data_validade,
                                        'quantidade_inicial': int(estoque_atual),
                                        'quantidade_atual': int(estoque_atual),
                                        'preco_custo_lote': preco_custo,
                                    }
                                )

                    except Exception as e:
                        logger.error(f"Erro ao processar linha {index} ({row.to_dict()}): {e}")

        except Exception as e:
            logger.exception(f"Erro geral ao processar arquivo: {e}")
            raise ValueError(f"Erro ao processar arquivo: {e}")

        return registros_criados, registros_atualizados

    # ------------------------------
    # Execução do formulário
    # ------------------------------
    def form_valid(self, form):
        empresa = self.get_empresa()
        if not empresa:
            messages.error(self.request, "Empresa não associada ao utilizador.")
            return self.form_invalid(form)

        arquivo = form.cleaned_data.get('arquivo')
        if not arquivo:
            messages.error(self.request, "Nenhum arquivo foi enviado.")
            return self.form_invalid(form)

        try:
            criados, atualizados = self.processar_arquivo(arquivo, empresa)
            messages.success(
                self.request,
                f"Importação concluída: {criados} produtos criados e {atualizados} atualizados."
            )
        except Exception as e:
            logger.exception(f"Erro ao importar produtos: {e}")
            messages.error(self.request, f"Erro ao importar produtos: {e}")

        return super().form_valid(form)


from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---- EXPORTAÇÃO XLSX ---- 
class ExportarProdutosExcelView(LoginRequiredMixin, View):
    def get(self, request):
        empresa = self.get_empresa()
        if not empresa:
            return HttpResponseBadRequest("Empresa não encontrada")

        produtos = Produto.objects.filter(empresa=empresa).select_related(
            'categoria', 'fornecedor', 'fabricante', 'taxa_iva'
        )

        def to_float(v):
            if v in (None, "", " ", "-"):
                return 0.0
            try:
                return float(str(v).replace(",", "."))
            except:
                return 0.0
        # ===================================
        # CÁLCULOS
        # ===================================
        total_investido = sum(to_float(p.preco_custo) * to_float(p.estoque_atual) for p in produtos)
        total_esperado = sum(to_float(p.preco_venda) * to_float(p.estoque_atual) for p in produtos)


        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"

        # ===================================
        # ESTILOS
        # ===================================
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        fill_header = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # ===================================
        # CABEÇALHO SUPERIOR COM OS TOTAIS
        # ===================================
        ws.append(["Total Investido (Custo × Estoque):", float(total_investido)])
        ws.append(["Total Esperado (Venda × Estoque):", float(total_esperado)])
        ws.append([])

        # Mesclar colunas A:B para títulos
        ws.merge_cells("A1:B1")
        ws.merge_cells("A2:B2")

        # Estilo dos totais
        ws["A1"].font = bold
        ws["A2"].font = bold
        ws["A1"].alignment = left
        ws["A2"].alignment = left

        # ===================================
        # TABELA PRINCIPAL
        # ===================================
        headers = [
            'nome_produto', 'codigo_barras', 'categoria', 'preco_custo', 'preco_venda',
            'Taxa IVA (%)',
            'nome_comercial', 'codigo_interno', 'fabricante', 'fornecedor',
            'estoque_atual', 'ativo'
        ]
        ws.append(headers)

        header_row = ws.max_row
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = bold
            cell.fill = fill_header
            cell.alignment = center
            cell.border = thin_border

        # ===================================
        # DADOS DOS PRODUTOS
        # ===================================
        for p in produtos:
            ws.append([
                p.nome_produto,
                p.codigo_barras,
                p.categoria.nome if p.categoria else '-',
                to_float(p.preco_custo),
                to_float(p.preco_venda),
                float(p.taxa_iva.tax_percentage) if p.taxa_iva else 0.00,
                p.nome_comercial,
                p.codigo_interno,
                p.fabricante.nome if p.fabricante else '-',
                p.fornecedor.nome_fantasia if p.fornecedor else '-',
                to_float(p.estoque_atual),
                'Ativo' if p.ativo else 'Inativo',
            ])

        # Aplicar bordas e alinhamento nas linhas de dados
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = center
                else:
                    cell.alignment = left

        # ===================================
        # AJUSTE DE LARGURA DAS COLUNAS
        # ===================================
        from openpyxl.utils import get_column_letter

        num_cols = ws.max_column

        for col_idx in range(1, num_cols + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))

            ws.column_dimensions[col_letter].width = max_length + 2


        # ===================================
        # EXPORTAÇÃO
        # ===================================
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="produtos_{empresa.nome}.xlsx"'
        return response

    def get_empresa(self):
        user = self.request.user
        if hasattr(user, 'funcionario') and user.funcionario.empresa:
            return user.funcionario.empresa
        return Empresa.objects.first() if user.is_superuser else None


# ---- EXPORTAÇÃO PDF ----
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from django.http import HttpResponse, HttpResponseBadRequest
import io, requests
from datetime import datetime

class ExportarProdutosPDFView(LoginRequiredMixin, View):
    def get(self, request):
        empresa = self.get_empresa()
        if not empresa:
            return HttpResponseBadRequest("Empresa não encontrada")

        produtos = Produto.objects.filter(empresa=empresa)

        buffer = io.BytesIO()
        pdf = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)

        elementos = []
        styles = getSampleStyleSheet()

        # ==============================
        # LOGOTIPO DA EMPRESA
        # ==============================
        # ==============================
        if empresa.foto and hasattr(empresa.foto, 'url'):
            try:
                resp = requests.get(empresa.foto.url)
                if resp.status_code == 200:
                    logo_stream = io.BytesIO(resp.content)
                    img = Image(logo_stream, width=4*cm, height=4*cm)
                    elementos.append(img)
                    elementos.append(Spacer(1, 12))
            except:
                pass

        # ==============================
        # CABEÇALHO
        # ==============================
        titulo = Paragraph(f"<b>Listagem de Produtos – {empresa.nome}</b>", styles["Title"])
        data_geracao = Paragraph(
            f"<font size=8>Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
            styles["Normal"]
        )
        elementos.append(titulo)
        elementos.append(data_geracao)
        elementos.append(Spacer(1, 20))

        # ==============================
        # TABELA DE PRODUTOS
        # ==============================

        tabela_dados = [
            ["Nº", "Produto", "Preço (AKZ)", "Taxa IVA", "Estoque"]
        ]

        for i, p in enumerate(produtos, start=1):
            iva_str = f"{p.taxa_iva.tax_percentage:.2f}%" if p.taxa_iva else "0.00%"
            tabela_dados.append([
                i,
                p.nome_produto,
                f"{p.preco_venda:.2f}",
                iva_str,
                p.estoque_atual
            ])

        tabela = Table(tabela_dados, colWidths=[1.5*cm, 8*cm, 3*cm, 2.5*cm, 2*cm])


        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),  # Cabeçalho
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),       # Cabeçalho
            ("ALIGN", (0, 1), (0, -1), "LEFT"),              # Coluna Nº centralizada
            ("ALIGN", (1, 1), (1, -1), "LEFT"),               # Coluna Produto alinhada à esquerda
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),            # Demais colunas centralizadas
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))


        elementos.append(tabela)

        # Construir PDF
        pdf.build(elementos)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="produtos_{empresa.nome}.pdf"'
        return response

    def get_empresa(self):
        user = self.request.user
        if hasattr(user, "funcionario") and user.funcionario.empresa:
            return user.funcionario.empresa
        return Empresa.objects.first() if user.is_superuser else None


